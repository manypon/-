"""
Yandex Maps Scanner — CustomTkinter GUI
====================================================

Запуск:
    python main.py

При первом запуске автоматически поставятся недостающие библиотеки:
customtkinter, playwright, pandas, openpyxl.

Используется уже установленный Chromium-based браузер (Яндекс.Браузер /
Chrome / Edge) — через поле "Браузер", либо скрипт попробует найти его сам.

Почему CustomTkinter, а не pywebview:
pywebview на Windows использует WebView2 через .NET-мост (pythonnet), а он
не всегда совместим со свежими версиями Python (в частности, ловит
рекурсию/COM-ошибки на Python 3.14). CustomTkinter — чистый Python поверх
tkinter, без .NET/COM-зависимостей, поэтому работает предсказуемо и
одинаково на любой актуальной версии Python.
"""

import re
import os
import sys
import time
import queue
import random
import threading
import subprocess
import importlib
from urllib.parse import quote


# ================= АВТОУСТАНОВКА ЗАВИСИМОСТЕЙ =================

def _ensure_package(pip_name, import_name=None):
    import_name = import_name or pip_name
    try:
        importlib.import_module(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pip_name])


def _ensure_dependencies():
    # В собранном .exe (PyInstaller) все зависимости уже вшиты внутрь —
    # sys.executable там указывает на сам .exe, а не на python.exe, так что
    # pip install через него не сработает и не нужен.
    if getattr(sys, "frozen", False):
        return
    for pip_name, import_name in [
        ("customtkinter", "customtkinter"),
        ("playwright", "playwright"),
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
    ]:
        _ensure_package(pip_name, import_name)


_ensure_dependencies()

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout


# ======================= СОЦСЕТИ / ФИЛЬТРЫ =======================

SOCIAL_PATTERNS = {
    "VK": re.compile(r'https?://(?:www\.)?vk\.com/[^\s"\'<>]+', re.I),
    "Telegram": re.compile(r'https?://(?:www\.)?t\.me/[^\s"\'<>]+', re.I),
    "Instagram": re.compile(r'https?://(?:www\.)?instagram\.com/[^\s"\'<>]+', re.I),
    "WhatsApp": re.compile(r'https?://(?:api\.|www\.)?wa\.me/[^\s"\'<>]+', re.I),
    "Facebook": re.compile(r'https?://(?:www\.)?facebook\.com/[^\s"\'<>]+', re.I),
    "OK.ru": re.compile(r'https?://(?:www\.)?ok\.ru/[^\s"\'<>]+', re.I),
}

IGNORE_DOMAINS = (
    "yandex.", "ya.ru", "vk.com", "t.me", "instagram.com",
    "facebook.com", "wa.me", "ok.ru", "adfox", "mc.yandex",
)

COLUMNS = [
    "Название", "Ссылка на карточку", "Адрес", "Телефон",
    "Есть сайт", "Сайт", "Рейтинг", "Кол-во отзывов",
    "VK", "Telegram", "Instagram", "WhatsApp", "Facebook", "OK.ru",
]

TABLE_COLUMNS = [
    "Название", "Адрес", "Телефон", "Сайт", "Рейтинг", "Отзывы",
    "VK", "Telegram", "Instagram", "WhatsApp", "Facebook", "OK.ru",
]


def find_browser(manual_path: str = "") -> str:
    """Ищет Chromium-based браузер: сначала ручной путь, потом Яндекс.Браузер,
    потом Edge, потом Chrome — в стандартных папках Windows."""
    if manual_path and os.path.isfile(manual_path):
        return manual_path

    local_appdata = os.environ.get("LOCALAPPDATA", "")
    program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    program_files_x86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")

    candidates = [
        os.path.join(local_appdata, "Yandex", "YandexBrowser", "Application", "browser.exe"),
        os.path.join(program_files, "Yandex", "YandexBrowser", "Application", "browser.exe"),
        os.path.join(program_files_x86, "Yandex", "YandexBrowser", "Application", "browser.exe"),
        os.path.join(program_files_x86, "Microsoft", "Edge", "Application", "msedge.exe"),
        os.path.join(program_files, "Microsoft", "Edge", "Application", "msedge.exe"),
        os.path.join(program_files, "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(program_files_x86, "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(local_appdata, "Google", "Chrome", "Application", "chrome.exe"),
    ]

    for path in candidates:
        if path and os.path.isfile(path):
            return path

    raise FileNotFoundError(
        "Не нашёл ни Яндекс.Браузер, ни Edge, ни Chrome автоматически. "
        "Впиши путь к browser.exe / msedge.exe / chrome.exe вручную в поле "
        "'Браузер' (узнать путь: в адресной строке браузера открой "
        "browser://version или edge://version или chrome://version)."
    )


def build_search_url(query: str, city: str) -> str:
    full_query = f"{query} {city}".strip()
    return f"https://yandex.ru/maps/?text={quote(full_query)}"


def is_captcha(page) -> bool:
    url = page.url
    if "showcaptcha" in url or "checkcaptcha" in url:
        return True
    content = page.content()
    return "captcha" in content.lower() and "yandex" in url.lower() and "maps" not in url.lower()


# ======================= ЛОГИКА ПАРСИНГА =======================

def collect_card_links(page, max_results, log_fn, stop_flag):
    links = set()
    try:
        page.wait_for_selector("a[href*='/maps/org/']", timeout=15000)
    except PWTimeout:
        log_fn("Не удалось дождаться списка результатов — проверь запрос или интернет.")
        return []

    scroller = page.locator("div[class*='scroll__container']").first
    if scroller.count() == 0:
        scroller = page.locator("body")

    stagnant_rounds = 0
    last_count = 0

    while len(links) < max_results and stagnant_rounds < 5:
        if stop_flag.is_set():
            log_fn("Остановлено пользователем.")
            break

        anchors = page.locator("a[href*='/maps/org/']")
        count = anchors.count()
        for i in range(count):
            href = anchors.nth(i).get_attribute("href")
            if href:
                if href.startswith("/"):
                    href = "https://yandex.ru" + href
                links.add(href.split("?")[0])

        if len(links) == last_count:
            stagnant_rounds += 1
        else:
            stagnant_rounds = 0
        last_count = len(links)

        try:
            scroller.hover()
            page.mouse.wheel(0, 1200)
        except Exception:
            pass

        time.sleep(random.uniform(0.8, 1.6))

        if is_captcha(page):
            log_fn("⚠️ Похоже, капча. Реши её в открытом окне браузера, жду 20 сек...")
            time.sleep(20)

    return list(links)[:max_results]


def extract_org_data(page, url, log_fn):
    data = {
        "Название": "", "Ссылка на карточку": url, "Адрес": "", "Телефон": "",
        "Есть сайт": "Нет", "Сайт": "", "Рейтинг": "", "Кол-во отзывов": "",
        "VK": "", "Telegram": "", "Instagram": "", "WhatsApp": "", "Facebook": "", "OK.ru": "",
    }
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except PWTimeout:
        return data

    if is_captcha(page):
        log_fn("⚠️ Похоже, капча на карточке. Жду 20 сек...")
        time.sleep(20)

    try:
        page.wait_for_selector("h1", timeout=10000)
    except PWTimeout:
        pass

    try:
        data["Название"] = page.locator("h1").first.inner_text().strip()
    except Exception:
        pass

    html = page.content()

    for key, pattern in SOCIAL_PATTERNS.items():
        m = pattern.search(html)
        if m:
            data[key] = m.group(0)

    try:
        contact_links = page.locator("a[href^='http']")
        n = contact_links.count()
        for i in range(n):
            href = contact_links.nth(i).get_attribute("href") or ""
            if href and not any(dom in href for dom in IGNORE_DOMAINS):
                data["Сайт"] = href
                data["Есть сайт"] = "Да"
                break
    except Exception:
        pass

    try:
        addr_el = page.locator("[class*='address']").first
        if addr_el.count() > 0:
            data["Адрес"] = addr_el.inner_text().strip().replace("\n", ", ")
    except Exception:
        pass

    tel_match = re.search(r'tel:([+\d\-\s()]{6,20})', html)
    if tel_match:
        data["Телефон"] = tel_match.group(1).strip()

    rating_match = re.search(r'"ratingValue"\s*:\s*"?([\d.,]+)"?', html)
    if rating_match:
        raw = rating_match.group(1).replace(",", ".")
        try:
            # у Яндекс.Карт рейтинг иногда приходит как float32
            # ("4.900000095367432") — округляем до 1 знака после запятой
            data["Рейтинг"] = f"{float(raw):.1f}"
        except ValueError:
            data["Рейтинг"] = raw

    reviews_match = re.search(r'"reviewCount"\s*:\s*"?(\d+)"?', html)
    if reviews_match:
        data["Кол-во отзывов"] = reviews_match.group(1)

    return data


def save_excel(results, path):
    df = pd.DataFrame(results, columns=COLUMNS)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Результаты")
        ws = writer.sheets["Результаты"]

        header_font = Font(name="Arial", bold=True, color="FFF3DA")
        header_fill = PatternFill("solid", fgColor="332612")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        widths = [24, 45, 27, 18, 11, 27, 10, 16, 31, 29, 38, 27, 34, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"


def run_parser(params, put_fn, stop_flag):
    """Работает в фоновом потоке. put_fn(msg_dict) кладёт события в очередь,
    которую основной поток вычитывает через after()."""
    try:
        browser_path = find_browser(params["browser_path"])
        put_fn({"type": "log", "text": f"Использую браузер: {browser_path}"})
    except FileNotFoundError as e:
        put_fn({"type": "log", "text": str(e)})
        put_fn({"type": "done", "success": False})
        return

    search_url = build_search_url(params["query"], params["city"])
    results = []

    def log_fn(text):
        put_fn({"type": "log", "text": text})

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=params["headless"], executable_path=browser_path)
            context = browser.new_context(locale="ru-RU")
            page = context.new_page()

            log_fn(f"Открываю поиск: {search_url}")
            page.goto(search_url, wait_until="domcontentloaded", timeout=30000)

            if is_captcha(page):
                log_fn("⚠️ Капча на старте. Реши её в окне браузера, жду 20 сек...")
                time.sleep(20)

            log_fn("Собираю ссылки на карточки организаций...")
            links = collect_card_links(page, params["max_results"], log_fn, stop_flag)
            log_fn(f"Найдено карточек: {len(links)}")

            for idx, link in enumerate(links, 1):
                if stop_flag.is_set():
                    log_fn("Остановлено пользователем.")
                    break
                log_fn(f"[{idx}/{len(links)}] {link}")
                org_data = extract_org_data(page, link, log_fn)
                results.append(org_data)
                put_fn({"type": "progress", "current": idx, "total": len(links)})
                put_fn({"type": "org", "data": org_data, "idx": idx, "total": len(links)})
                time.sleep(random.uniform(1.5, 3.5))

            browser.close()
    except Exception as e:
        log_fn(f"Ошибка: {e}")
        put_fn({"type": "done", "success": False})
        return

    if not results:
        log_fn("Не удалось собрать данные.")
        put_fn({"type": "done", "success": False})
        return

    try:
        save_excel(results, params["output_file"])
        log_fn(f"✅ Готово! Сохранено {len(results)} строк в: {params['output_file']}")
        put_fn({"type": "done", "success": True, "path": params["output_file"]})
    except Exception as e:
        log_fn(f"Не удалось сохранить Excel: {e}")
        put_fn({"type": "done", "success": False})


# ======================= ЦВЕТОВАЯ СХЕМА =======================

ctk.set_appearance_mode("dark")

BG_ROOT = "#0D0F0A"
BG_PANEL = "#14160F"
BG_PANEL_2 = "#191C13"
BORDER = "#33362A"
TEXT = "#E7E3D3"
MUTED = "#8B8B78"
AMBER = "#E8A33D"
AMBER_DIM = "#7A5A28"
GREEN = "#8FFF6B"
DANGER = "#FF5C5C"

FONT_TITLE = ("Arial", 20, "bold")
FONT_SUB = ("Consolas", 11)
FONT_LABEL = ("Consolas", 11)
FONT_MONO = ("Consolas", 12)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Yandex Maps · Scanner")
        self.geometry("1180x820")
        self.minsize(980, 700)
        self.configure(fg_color=BG_ROOT)

        self.msg_queue = queue.Queue()
        self.stop_flag = threading.Event()
        self.worker_thread = None
        self.last_output_path = ""
        self._found = 0
        self._with_site = 0
        self._with_social = 0
        self._rating_sum = 0.0
        self._rating_count = 0

        self._build_header()
        self._build_tabs()
        self._build_main_tab()
        self._build_table_tab()
        self._show_tab("main")

        self.after(150, self._poll_queue)

    # ---------- HEADER ----------

    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=24, pady=(22, 8))

        left = ctk.CTkFrame(header, fg_color="transparent")
        left.pack(side="left")
        ctk.CTkLabel(left, text="🗺️  Yandex Maps · Scanner", font=FONT_TITLE, text_color=TEXT).pack(anchor="w")
        ctk.CTkLabel(left, text="СБОР ДАННЫХ ОБ ОРГАНИЗАЦИЯХ", font=FONT_SUB, text_color=MUTED).pack(anchor="w", pady=(2, 0))

        self.status_pill = ctk.CTkLabel(
            header, text="●  Готов к запуску", font=FONT_SUB, text_color=MUTED,
            fg_color=BG_PANEL, corner_radius=6, padx=14, pady=6,
        )
        self.status_pill.pack(side="right")

        hazard = ctk.CTkFrame(self, fg_color=AMBER_DIM, height=3, corner_radius=0)
        hazard.pack(fill="x", padx=24, pady=(4, 14))

    # ---------- TABS ----------

    def _build_tabs(self):
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=24, pady=(0, 12))

        self.tab_main_btn = ctk.CTkButton(
            row, text="ОСНОВНОЕ", width=140, height=34, corner_radius=8,
            fg_color=AMBER, text_color="#1A1204", hover_color="#f4b352",
            font=("Arial", 12, "bold"), command=lambda: self._show_tab("main"),
        )
        self.tab_main_btn.pack(side="left")

        self.tab_table_btn = ctk.CTkButton(
            row, text="ТАБЛИЦА", width=140, height=34, corner_radius=8,
            fg_color=BG_PANEL_2, text_color=MUTED, hover_color="#22251A",
            font=("Arial", 12, "bold"), command=lambda: self._show_tab("table"),
        )
        self.tab_table_btn.pack(side="left", padx=(8, 0))

    def _show_tab(self, name):
        if name == "main":
            self.tab_main_btn.configure(fg_color=AMBER, text_color="#1A1204")
            self.tab_table_btn.configure(fg_color=BG_PANEL_2, text_color=MUTED)
            self.table_tab.pack_forget()
            self.main_tab.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        else:
            self.tab_table_btn.configure(fg_color=AMBER, text_color="#1A1204")
            self.tab_main_btn.configure(fg_color=BG_PANEL_2, text_color=MUTED)
            self.main_tab.pack_forget()
            self.table_tab.pack(fill="both", expand=True, padx=24, pady=(0, 16))

    # ---------- ВКЛАДКА "ОСНОВНОЕ" ----------

    def _build_main_tab(self):
        self.main_tab = ctk.CTkFrame(self, fg_color="transparent")

        grid = ctk.CTkFrame(self.main_tab, fg_color="transparent")
        grid.pack(fill="both", expand=True)
        grid.grid_columnconfigure(0, weight=6)
        grid.grid_columnconfigure(1, weight=5)
        grid.grid_rowconfigure(0, weight=1)

        self._build_form_panel(grid)
        self._build_feed_panel(grid)
        self._build_log(self.main_tab)

    def _panel(self, parent, title):
        card = ctk.CTkFrame(parent, fg_color=BG_PANEL, corner_radius=4, border_width=1, border_color=BORDER)
        ctk.CTkLabel(card, text=title.upper(), font=("Arial", 11, "bold"), text_color=AMBER).pack(
            anchor="w", padx=18, pady=(16, 12)
        )
        return card

    def _build_form_panel(self, grid):
        card = self._panel(grid, "Параметры поиска")
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=18, pady=(0, 18))
        inner.grid_columnconfigure(0, weight=1)
        inner.grid_columnconfigure(1, weight=1)

        self.query_var = ctk.StringVar(value="кофейня")
        self.city_var = ctk.StringVar(value="Москва")
        self._labeled_entry(inner, "Запрос", self.query_var, row=0, col=0)
        self._labeled_entry(inner, "Город", self.city_var, row=0, col=1)

        self.max_var = ctk.StringVar(value="30")
        self._labeled_entry(inner, "Лимит карточек", self.max_var, row=1, col=0)

        switch_wrap = ctk.CTkFrame(inner, fg_color="transparent")
        switch_wrap.grid(row=1, column=1, sticky="ew", padx=6, pady=(16, 0))
        self.headless_var = ctk.BooleanVar(value=False)
        ctk.CTkLabel(switch_wrap, text="Headless-режим", font=FONT_LABEL, text_color=MUTED).pack(anchor="w")
        ctk.CTkSwitch(
            switch_wrap, text="", variable=self.headless_var, width=38,
            progress_color=AMBER_DIM, button_color=AMBER, fg_color=BG_PANEL_2,
        ).pack(anchor="w", pady=(6, 0))

        ctk.CTkFrame(inner, fg_color=BORDER, height=1).grid(row=2, column=0, columnspan=2, sticky="ew", pady=18)

        ctk.CTkLabel(inner, text="Папка для сохранения", font=FONT_LABEL, text_color=MUTED).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=6
        )
        folder_row = ctk.CTkFrame(inner, fg_color="transparent")
        folder_row.grid(row=4, column=0, columnspan=2, sticky="ew", padx=6, pady=(4, 14))
        folder_row.grid_columnconfigure(0, weight=1)
        self.folder_var = ctk.StringVar(value=os.getcwd())
        ctk.CTkEntry(
            folder_row, textvariable=self.folder_var, height=36, corner_radius=4,
            fg_color=BG_PANEL_2, border_color=BORDER, text_color=TEXT,
        ).grid(row=0, column=0, sticky="ew")
        ctk.CTkButton(
            folder_row, text="Обзор", width=84, height=36, corner_radius=4,
            fg_color=BG_PANEL_2, hover_color="#22251A", text_color=MUTED,
            border_width=1, border_color=BORDER, command=self._browse_folder,
        ).grid(row=0, column=1, padx=(8, 0))

        self.filename_var = ctk.StringVar(value="yandex_maps_result")
        self._labeled_entry(inner, "Имя файла (без .xlsx)", self.filename_var, row=5, col=0, span=2)

        ctk.CTkLabel(inner, text="Браузер (необязательно)", font=FONT_LABEL, text_color=MUTED).grid(
            row=6, column=0, columnspan=2, sticky="w", padx=6, pady=(14, 0)
        )
        browser_row = ctk.CTkFrame(inner, fg_color="transparent")
        browser_row.grid(row=7, column=0, columnspan=2, sticky="ew", padx=6, pady=(4, 0))
        browser_row.grid_columnconfigure(0, weight=1)
        self.browser_var = ctk.StringVar(value="")
        try:
            self.browser_var.set(find_browser(""))
        except FileNotFoundError:
            pass
        ctk.CTkEntry(
            browser_row, textvariable=self.browser_var, height=36, corner_radius=4,
            fg_color=BG_PANEL_2, border_color=BORDER, text_color=TEXT,
            placeholder_text="Найдётся автоматически: Яндекс.Браузер / Edge / Chrome",
        ).grid(row=0, column=0, sticky="ew")
        ctk.CTkButton(
            browser_row, text="Обзор", width=84, height=36, corner_radius=4,
            fg_color=BG_PANEL_2, hover_color="#22251A", text_color=MUTED,
            border_width=1, border_color=BORDER, command=self._browse_browser,
        ).grid(row=0, column=1, padx=(8, 0))

        actions = ctk.CTkFrame(inner, fg_color="transparent")
        actions.grid(row=8, column=0, columnspan=2, sticky="ew", pady=(20, 0))
        self.start_btn = ctk.CTkButton(
            actions, text="▶  Начать", height=42, corner_radius=6,
            fg_color=AMBER, hover_color="#f4b352", text_color="#1A1204",
            font=("Arial", 13, "bold"), command=self._start,
        )
        self.start_btn.pack(side="left", fill="x", expand=True)
        self.stop_btn = ctk.CTkButton(
            actions, text="■  Стоп", height=42, corner_radius=6,
            fg_color="transparent", hover_color="#2A1414", text_color=DANGER,
            border_width=1, border_color="#4A2020",
            font=("Arial", 13, "bold"), command=self._stop, state="disabled",
        )
        self.stop_btn.pack(side="left", fill="x", expand=True, padx=(8, 0))

    def _labeled_entry(self, parent, label, var, row, col, span=1):
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.grid(row=row, column=col, columnspan=span, sticky="ew", padx=6)
        ctk.CTkLabel(wrap, text=label, font=FONT_LABEL, text_color=MUTED).pack(anchor="w")
        ctk.CTkEntry(
            wrap, textvariable=var, height=36, corner_radius=4,
            fg_color=BG_PANEL_2, border_color=BORDER, text_color=TEXT,
        ).pack(fill="x", pady=(4, 10))

    def _build_feed_panel(self, grid):
        card = self._panel(grid, "Обнаруженные объекты")
        card.grid(row=0, column=1, sticky="nsew")
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=18, pady=(0, 18))

        stats = ctk.CTkFrame(inner, fg_color="transparent")
        stats.pack(fill="x", pady=(0, 12))
        for i in range(4):
            stats.grid_columnconfigure(i, weight=1)

        self.stat_found = self._stat_tile(stats, "0/0", "Найдено", 0, AMBER)
        self.stat_site = self._stat_tile(stats, "0", "С сайтом", 1, GREEN)
        self.stat_social = self._stat_tile(stats, "0", "С соцсетями", 2, GREEN)
        self.stat_rating = self._stat_tile(stats, "—", "Ср. рейтинг", 3, AMBER)

        self.progress = ctk.CTkProgressBar(inner, height=8, corner_radius=4, progress_color=AMBER, fg_color=BG_PANEL_2)
        self.progress.set(0)
        self.progress.pack(fill="x", pady=(0, 12))

        self.feed_scroll = ctk.CTkScrollableFrame(inner, fg_color="transparent", height=320)
        self.feed_scroll.pack(fill="both", expand=True)
        self.feed_empty_label = ctk.CTkLabel(self.feed_scroll, text="Пока ничего не найдено", font=FONT_LABEL, text_color=MUTED)
        self.feed_empty_label.pack(pady=20)

        legend = ctk.CTkFrame(inner, fg_color="transparent")
        legend.pack(fill="x", pady=(10, 0))
        self._legend_item(legend, "Есть сайт", AMBER)
        self._legend_item(legend, "Есть соцсети", GREEN)
        self._legend_item(legend, "Не найдено", BORDER)

    def _stat_tile(self, parent, value, label, col, color):
        tile = ctk.CTkFrame(parent, fg_color=BG_PANEL_2, corner_radius=4, border_width=1, border_color=BORDER)
        tile.grid(row=0, column=col, sticky="nsew", padx=4)
        val_lbl = ctk.CTkLabel(tile, text=value, font=("Arial", 18, "bold"), text_color=color)
        val_lbl.pack(pady=(10, 0))
        ctk.CTkLabel(tile, text=label.upper(), font=("Consolas", 9), text_color=MUTED).pack(pady=(0, 10))
        return val_lbl

    def _legend_item(self, parent, text, color):
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.pack(side="left", padx=(0, 16))
        dot = ctk.CTkFrame(wrap, fg_color=color, width=8, height=8, corner_radius=2)
        dot.pack(side="left", padx=(0, 5))
        ctk.CTkLabel(wrap, text=text, font=("Consolas", 9), text_color=MUTED).pack(side="left")

    def _build_log(self, parent):
        card = ctk.CTkFrame(parent, fg_color="#0A0B07", corner_radius=4, border_width=1, border_color=BORDER)
        card.pack(fill="x", pady=(14, 0))
        self.log_box = ctk.CTkTextbox(
            card, height=170, corner_radius=0, fg_color="#0A0B07",
            text_color=TEXT, font=FONT_MONO, wrap="word",
        )
        self.log_box.pack(fill="both", expand=True, padx=12, pady=10)
        self.log_box.configure(state="disabled")
        self.log_box.tag_config("warn", foreground=AMBER)
        self.log_box.tag_config("ok", foreground=GREEN)
        self.log_box.tag_config("err", foreground=DANGER)
        self.log_box.tag_config("muted", foreground=MUTED)

    # ---------- ВКЛАДКА "ТАБЛИЦА" ----------

    def _build_table_tab(self):
        self.table_tab = ctk.CTkFrame(self, fg_color="transparent")

        toolbar = ctk.CTkFrame(self.table_tab, fg_color="transparent")
        toolbar.pack(fill="x", pady=(0, 10))
        self.table_count_label = ctk.CTkLabel(
            toolbar, text="0 из 0 · обновляется в реальном времени", font=FONT_LABEL, text_color=MUTED,
        )
        self.table_count_label.pack(side="left")
        ctk.CTkButton(
            toolbar, text="⬇  Открыть Excel-файл", height=34, corner_radius=4,
            fg_color="transparent", hover_color="#2A2312", text_color=AMBER,
            border_width=1, border_color=AMBER_DIM, command=self._open_output_file,
        ).pack(side="right")

        card = ctk.CTkFrame(self.table_tab, fg_color=BG_PANEL, corner_radius=4, border_width=1, border_color=BORDER)
        card.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Dark.Treeview", background=BG_PANEL_2, fieldbackground=BG_PANEL_2,
            foreground=TEXT, rowheight=26, borderwidth=0, font=("Consolas", 10),
        )
        style.configure(
            "Dark.Treeview.Heading", background="#20240F", foreground=AMBER,
            font=("Arial", 9, "bold"), borderwidth=0,
        )
        style.map("Dark.Treeview", background=[("selected", "#332612")])

        tree_wrap = ctk.CTkFrame(card, fg_color="transparent")
        tree_wrap.pack(fill="both", expand=True, padx=12, pady=12)

        self.tree = ttk.Treeview(
            tree_wrap, columns=TABLE_COLUMNS, show="headings", style="Dark.Treeview",
        )
        for col in TABLE_COLUMNS:
            self.tree.heading(col, text=col.upper())
            self.tree.column(col, width=130, anchor="w")

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_wrap.grid_rowconfigure(0, weight=1)
        tree_wrap.grid_columnconfigure(0, weight=1)

    # ---------- ДЕЙСТВИЯ ----------

    def _browse_folder(self):
        path = filedialog.askdirectory(initialdir=self.folder_var.get() or os.getcwd())
        if path:
            self.folder_var.set(path)

    def _browse_browser(self):
        path = filedialog.askopenfilename(filetypes=[("Исполняемый файл", "*.exe"), ("Все файлы", "*.*")])
        if path:
            self.browser_var.set(path)

    def _open_output_file(self):
        path = self.last_output_path
        if not path or not os.path.isfile(path):
            messagebox.showinfo("Файл не найден", "Сначала запусти парсинг и дождись сохранения файла.")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # noqa
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

    def _put(self, msg):
        self.msg_queue.put(msg)

    def _reset_counters(self):
        self._found = 0
        self._with_site = 0
        self._with_social = 0
        self._rating_sum = 0.0
        self._rating_count = 0
        for widget in self.feed_scroll.winfo_children():
            widget.destroy()
        self.feed_empty_label = ctk.CTkLabel(self.feed_scroll, text="Пока ничего не найдено", font=FONT_LABEL, text_color=MUTED)
        self.feed_empty_label.pack(pady=20)
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.stat_found.configure(text="0/0")
        self.stat_site.configure(text="0")
        self.stat_social.configure(text="0")
        self.stat_rating.configure(text="—")
        self.progress.set(0)
        self.table_count_label.configure(text="0 из 0 · обновляется в реальном времени")
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _start(self):
        query = self.query_var.get().strip()
        if not query:
            messagebox.showwarning("Внимание", "Введи, что искать.")
            return
        try:
            max_results = int(self.max_var.get())
            if max_results <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Внимание", "Лимит карточек должен быть положительным числом.")
            return

        filename = self.filename_var.get().strip()
        if not filename:
            messagebox.showwarning("Внимание", "Введи имя файла.")
            return
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        folder = self.folder_var.get().strip() or os.getcwd()
        if not os.path.isdir(folder):
            messagebox.showwarning("Внимание", "Указанная папка не существует.")
            return

        output_path = os.path.join(folder, filename)
        self.last_output_path = output_path

        params = {
            "query": query,
            "city": self.city_var.get().strip(),
            "max_results": max_results,
            "output_file": output_path,
            "browser_path": self.browser_var.get().strip(),
            "headless": self.headless_var.get(),
        }

        self._reset_counters()
        self.stop_flag.clear()
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.status_pill.configure(text="●  Сканирование", text_color=GREEN)
        self._log(f"Запускаю поиск: '{params['query']}' в городе '{params['city']}'...", "muted")

        self.worker_thread = threading.Thread(
            target=run_parser, args=(params, self._put, self.stop_flag), daemon=True,
        )
        self.worker_thread.start()

    def _stop(self):
        self.stop_flag.set()
        self._log("Останавливаю после текущей карточки...", "muted")
        self.stop_btn.configure(state="disabled")

    # ---------- ОБРАБОТКА ОЧЕРЕДИ (только из главного потока) ----------

    def _log(self, text, tag=None):
        if tag is None:
            if "⚠️" in text:
                tag = "warn"
            elif "✅" in text:
                tag = "ok"
            elif "ошибка" in text.lower():
                tag = "err"
            elif text.startswith("["):
                tag = "muted"
        self.log_box.configure(state="normal")
        self.log_box.insert("end", "> " + text + "\n", tag or "")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _poll_queue(self):
        while not self.msg_queue.empty():
            msg = self.msg_queue.get_nowait()
            mtype = msg.get("type")
            if mtype == "log":
                self._log(msg["text"])
            elif mtype == "progress":
                self.progress.set(msg["current"] / max(msg["total"], 1))
            elif mtype == "org":
                self._add_org(msg["data"], msg["idx"], msg["total"])
            elif mtype == "done":
                self._on_done(msg.get("success", False), msg.get("path"))
        self.after(150, self._poll_queue)

    def _add_org(self, data, idx, total):
        self._found += 1
        has_site = data.get("Есть сайт") == "Да"
        has_social = any(data.get(k) for k in ("VK", "Telegram", "Instagram", "WhatsApp", "Facebook", "OK.ru"))
        if has_site:
            self._with_site += 1
        if has_social:
            self._with_social += 1

        try:
            rating = float(str(data.get("Рейтинг", "")).replace(",", "."))
            self._rating_sum += rating
            self._rating_count += 1
        except ValueError:
            pass

        self.stat_found.configure(text=f"{self._found}/{total}")
        self.stat_site.configure(text=str(self._with_site))
        self.stat_social.configure(text=str(self._with_social))
        if self._rating_count:
            self.stat_rating.configure(text=f"{self._rating_sum / self._rating_count:.1f}")
        self.table_count_label.configure(text=f"{self._found} из {total} · обновляется в реальном времени")

        if self.feed_empty_label:
            self.feed_empty_label.destroy()
            self.feed_empty_label = None

        card = ctk.CTkFrame(self.feed_scroll, fg_color=BG_PANEL_2, corner_radius=4, border_width=1,
                             border_color=(AMBER_DIM if idx == total or True else BORDER))
        card.pack(fill="x", pady=4)
        row = ctk.CTkFrame(card, fg_color="transparent")
        row.pack(fill="x", padx=10, pady=8)

        ctk.CTkLabel(row, text=str(idx), font=("Consolas", 10), text_color=MUTED, width=24).pack(side="left")

        mid = ctk.CTkFrame(row, fg_color="transparent")
        mid.pack(side="left", fill="x", expand=True, padx=8)
        ctk.CTkLabel(mid, text=data.get("Название") or "Без названия", font=("Consolas", 11), text_color=TEXT, anchor="w").pack(fill="x")
        ctk.CTkLabel(mid, text=data.get("Адрес") or "—", font=("Consolas", 9), text_color=MUTED, anchor="w").pack(fill="x")

        rating_text = f"★ {data['Рейтинг']}" if data.get("Рейтинг") else "—"
        ctk.CTkLabel(row, text=rating_text, font=("Arial", 11, "bold"), text_color=AMBER, width=44).pack(side="left")

        icons = ctk.CTkFrame(row, fg_color="transparent")
        icons.pack(side="left", padx=(6, 0))
        self._ico(icons, "W", has_site, AMBER)
        self._ico(icons, "S", has_social, GREEN)

        # таблица
        self.tree.insert("", "end", values=(
            data.get("Название") or "—",
            data.get("Адрес") or "—",
            data.get("Телефон") or "—",
            data.get("Сайт") or ("нет" if not has_site else ""),
            data.get("Рейтинг") or "—",
            data.get("Кол-во отзывов") or "—",
            data.get("VK") or "—",
            data.get("Telegram") or "—",
            data.get("Instagram") or "—",
            data.get("WhatsApp") or "—",
            data.get("Facebook") or "—",
            data.get("OK.ru") or "—",
        ))

    def _ico(self, parent, letter, found, color):
        fg = color if found else "#0D0E09"
        border = color if found else BORDER
        text_color = "#0D0F0A" if found else "#4A4C3E"
        box = ctk.CTkLabel(
            parent, text=letter, width=22, height=22, corner_radius=0,
            fg_color=fg, text_color=text_color, font=("Consolas", 9, "bold"),
        )
        box.pack(side="left", padx=2)

    def _on_done(self, success, path=None):
        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        if success:
            self.status_pill.configure(text="●  Готово ✅", text_color=GREEN)
            messagebox.showinfo("Готово", f"Файл сохранён:\n{path or self.last_output_path}")
        else:
            self.status_pill.configure(text="●  Завершено с ошибкой", text_color=DANGER)


if __name__ == "__main__":
    app = App()
    app.mainloop()
